from __future__ import annotations

import json
from pathlib import Path

from trace_tasks.core.taxonomy import TASK_TAXONOMY
from trace_tasks.review.app.store import ReviewStore
from trace_tasks.review.export import export_review_report
from tests.test_review_app_public import _DIGEST_A, _materialize_one


def _review_state(path: Path) -> tuple[str, ReviewStore]:
    task_id = next(iter(sorted(TASK_TAXONOMY)))
    store = ReviewStore(path)
    store.update_task_audit(
        task_id,
        _DIGEST_A,
        values={"prompt": True, "answer": True},
        notes="checked",
        updated_by="reviewer",
    )
    issue = store.create_issue(
        title="Example issue",
        body="Example body",
        task_id=task_id,
        recipe_digest=_DIGEST_A,
        category="prompt",
    )
    store.add_comment(issue["id"], body="Follow up")
    store.set_asset_review("sample-asset", kind="illustrations", decision="approve")
    return task_id, store


def _write_calibration(review_root: Path, task_id: str) -> None:
    path = review_root.parent / "calibration" / "run.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            {
                "schema": "trace-review-calibration-v1",
                "created_at": "2026-07-21T00:00:00+00:00",
                "model": "local/model",
                "selection": {"task_ids": [task_id], "sample_count": 1},
                "summary": {"mean_answer_reward": 0.5, "rollout_count": 2},
                "results": [{"task_id": task_id, "response": "not exported"}],
            }
        ),
        encoding="utf-8",
    )


def test_json_and_jsonl_exports_are_portable(tmp_path: Path) -> None:
    review_root = tmp_path / "review" / "task-reviews"
    database = tmp_path / "review" / "feedback" / "state.sqlite"
    _materialize_one(review_root)
    task_id, _ = _review_state(database)
    _write_calibration(review_root, task_id)

    json_path = export_review_report(
        review_root=review_root,
        database_path=database,
        output_path=tmp_path / "review.json",
    )
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["schema"] == "trace-review-report-v1"
    assert payload["summary"]["task_count"] == len(TASK_TAXONOMY)
    assert payload["issues"][0]["comments"][0]["body"] == "Follow up"
    assert payload["task_audits"][0]["task_id"] == task_id
    task = next(item for item in payload["tasks"] if item["task_id"] == task_id)
    assert task["recipe_digest"] == _DIGEST_A
    assert task["source_provenance"]["revision"] == "0123456789abcdef"
    assert payload["calibration_summaries"][0]["model"] == "local/model"
    assert "not exported" not in json_path.read_text(encoding="utf-8")
    assert str(tmp_path) not in json_path.read_text(encoding="utf-8")

    jsonl_path = export_review_report(
        review_root=review_root,
        database_path=database,
        output_path=tmp_path / "review.jsonl",
        output_format="jsonl",
    )
    records = [json.loads(line) for line in jsonl_path.read_text().splitlines()]
    assert records[0]["record_type"] == "summary"
    assert {record["record_type"] for record in records} >= {
        "task",
        "issue",
        "task_audit",
        "asset_review",
    }


def test_optional_workbook_export(tmp_path: Path) -> None:
    database = tmp_path / "state.sqlite"
    _review_state(database)
    path = export_review_report(
        review_root=tmp_path / "task-reviews",
        database_path=database,
        output_path=tmp_path / "review.xlsx",
        output_format="xlsx",
    )
    assert path.is_file()
    assert path.read_bytes().startswith(b"PK")


def test_workbook_neutralizes_formula_prefixed_reviewer_strings(
    tmp_path: Path,
) -> None:
    from openpyxl import load_workbook

    database = tmp_path / "state.sqlite"
    task_id = next(iter(sorted(TASK_TAXONOMY)))
    store = ReviewStore(database)
    store.update_task_audit(
        task_id,
        _DIGEST_A,
        values={"prompt": True},
        notes='=HYPERLINK("https://invalid")',
        updated_by="+reviewer",
    )
    issue = store.create_issue(
        title="=1+1",
        body="+cmd",
        author="@reviewer",
        task_id=task_id,
        recipe_digest=_DIGEST_A,
    )
    store.add_comment(issue["id"], body="-danger", author="=author")
    store.set_asset_review(
        "asset", kind="illustrations", decision="improve", notes="@formula"
    )
    path = export_review_report(
        review_root=tmp_path / "task-reviews",
        database_path=database,
        output_path=tmp_path / "safe.xlsx",
        output_format="xlsx",
    )
    workbook = load_workbook(path, data_only=False)
    reviewer_values = {
        value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if isinstance(value, str) and value.startswith("'")
    }
    assert {
        '\'=HYPERLINK("https://invalid")',
        "'+reviewer",
        "'=1+1",
        "'+cmd",
        "'@reviewer",
        "'@formula",
    }.issubset(reviewer_values)


def test_jsonl_preserves_sanitized_index_and_calibration_errors(
    tmp_path: Path,
) -> None:
    review_root = tmp_path / "review" / "task-reviews"
    task_id, domain, scene_id = _materialize_one(review_root)
    manifest = review_root / domain / scene_id / task_id / "manifest.json"
    manifest.write_text("{", encoding="utf-8")
    calibration = review_root.parent / "calibration"
    calibration.mkdir(parents=True)
    (calibration / "broken.json").write_text("{", encoding="utf-8")
    output = export_review_report(
        review_root=review_root,
        database_path=tmp_path / "state.sqlite",
        output_path=tmp_path / "errors.jsonl",
        output_format="jsonl",
    )
    records = [json.loads(line) for line in output.read_text().splitlines()]
    types = {record["record_type"] for record in records}
    assert {"index_error", "calibration_error"}.issubset(types)
    assert str(tmp_path) not in output.read_text(encoding="utf-8")
