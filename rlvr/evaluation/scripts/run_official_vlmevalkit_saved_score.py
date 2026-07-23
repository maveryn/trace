#!/usr/bin/env python3
"""Run VLMEvalKit's evaluator against an existing prediction workbook.

The wrapper owns staging and provenance only. Benchmark parsing, judge calls,
scoring, and metric aggregation stay inside ``dataset.evaluate``.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import inspect
import json
import math
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any, Callable

REPO_ROOT = Path(__file__).resolve().parents[3]
SCRIPTS_ROOT = Path(__file__).resolve().parent
EXTENSIONS_ROOT = Path(__file__).resolve().parents[1] / "vlmevalkit_extensions"
DEFAULT_VLMEVAL_ROOT = REPO_ROOT / "external" / "VLMEvalKit"
PINNED_VLMEVALKIT_COMMIT = "a8b12bf1c3737a33fc1de967c202f9c592b22e86"
CONTRACT = "vlmevalkit-dataset-evaluate-saved-response-v2"
CHARTQAPRO_ADAPTER_CONTRACT = "chartqapro-official-final-answer-v2"
PHYX_OPTION_ADAPTER_CONTRACT = "phyx-deterministic-option-v1"
TREEBENCH_OPTION_ADAPTER_CONTRACT = "treebench-explicit-final-boxed-option-v1"
REALWORLDQA_OPTION_ADAPTER_CONTRACT = "realworldqa-deterministic-final-option-v1"
EVALUATOR_INPUT_FILTER_CONTRACT = "official-evaluator-drop-queue-columns-v1"
QUEUE_ONLY_PREDICTION_COLUMNS = ("request_hash", "source_row_hash")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_default(value: Any) -> Any:
    try:
        import numpy as np

        if isinstance(value, np.integer):
            return int(value)
        if isinstance(value, np.floating):
            return float(value)
        if isinstance(value, np.bool_):
            return bool(value)
        if isinstance(value, np.ndarray):
            return value.tolist()
    except ImportError:
        pass
    if isinstance(value, Path):
        return str(value)
    return str(value)


def _write_json(path: Path, value: Any) -> None:
    temporary = path.with_suffix(path.suffix + f".tmp.{os.getpid()}")
    temporary.write_text(
        json.dumps(value, indent=2, ensure_ascii=False, default=_json_default) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def _parse_json_object(raw: str, argument: str) -> dict[str, Any]:
    try:
        value = json.loads(raw)
    except json.JSONDecodeError as error:
        raise ValueError(f"{argument} must be a valid JSON object: {error}") from error
    if not isinstance(value, dict):
        raise ValueError(f"{argument} must decode to a JSON object")
    return value


def _redact(mapping: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in mapping.items():
        lowered = str(key).lower().replace("-", "_")
        sensitive = (
            lowered in {"token", "secret", "password", "api_key", "authorization"}
            or lowered.endswith(("_token", "_secret", "_password", "_api_key"))
            or lowered.startswith(
                ("token_", "secret_", "password_", "api_key_", "authorization_")
            )
        )
        if sensitive:
            result[str(key)] = "<redacted>"
        elif isinstance(value, dict):
            result[str(key)] = _redact(value)
        else:
            result[str(key)] = value
    return result


def _git(root: Path, *args: str) -> str:
    return subprocess.run(
        ["git", "-C", str(root), *args],
        check=True,
        capture_output=True,
        text=True,
    ).stdout.strip()


def _normalize_result(result: Any) -> dict[str, Any]:
    import pandas as pd

    if isinstance(result, pd.DataFrame):
        return {"table": result.to_dict(orient="records")}
    if isinstance(result, dict):
        return json.loads(json.dumps(result, default=_json_default))
    raise TypeError(
        "VLMEvalKit dataset.evaluate must return a dict or DataFrame, "
        f"got {type(result).__name__}"
    )


def _primary_score(
    dataset: Any,
    result: Any,
    flatten_metrics: Callable[[Any], dict[str, Any]],
    requested_metric: str | None,
    value_scale: str,
) -> tuple[float, dict[str, Any]]:
    flattened = flatten_metrics(result)
    official = dataset.report_primary_metric(flattened)
    if not isinstance(official, dict):
        raise TypeError(
            "VLMEvalKit dataset.report_primary_metric did not return a dict"
        )

    if (
        requested_metric == "Score (Strict)"
        and getattr(dataset, "dataset_name", None) == "WeMath_COT"
        and hasattr(result, "columns")
        and requested_metric in result.columns
        and len(result) > 0
    ):
        key, raw = requested_metric, result.iloc[0][requested_metric]
    elif requested_metric:
        source = official if requested_metric in official else flattened
        if requested_metric not in source:
            available = sorted({*map(str, official), *map(str, flattened)})
            raise KeyError(
                f"Unknown primary metric {requested_metric!r}; available: {available}"
            )
        key, raw = requested_metric, source[requested_metric]
    elif len(official) == 1:
        key, raw = next(iter(official.items()))
    else:
        raise ValueError(
            "VLMEvalKit did not report exactly one primary metric; "
            f"pass --primary-metric explicitly. Reported: {official!r}"
        )

    percent_text = isinstance(raw, str) and raw.strip().endswith("%")
    value = float(raw.strip()[:-1] if percent_text else raw)
    if not math.isfinite(value):
        raise ValueError(f"Primary metric {key!r} is not finite: {raw!r}")
    if value_scale == "fraction" or (
        value_scale == "auto" and not percent_text and abs(value) <= 1.0
    ):
        score = value * 100.0
    else:
        score = value
    return score, {
        "key": str(key),
        "raw_value": value,
        "raw_display": str(raw) if percent_text else None,
        "value_scale": value_scale,
        "official_report": official,
    }


def _stage_prediction(
    source: Path,
    output_dir: Path,
    dataset_name: str,
    *,
    replace: bool = False,
) -> Path:
    name = str(dataset_name).replace("/", "_").replace(os.sep, "_")
    target = output_dir / f"{name}_predictions.xlsx"
    if source.resolve() == target.resolve():
        return target
    if replace:
        shutil.copy2(source, target)
    elif target.exists():
        if _sha256(target) != _sha256(source):
            raise FileExistsError(
                f"Different staged prediction already exists at {target}"
            )
    else:
        shutil.copy2(source, target)
    return target


def _xlsx_header(path: Path) -> list[Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook.active
        return [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    finally:
        workbook.close()


def _filter_evaluator_input(prediction: Path) -> dict[str, Any]:
    """Remove queue-only identities from the copy passed to VLMEvalKit."""

    from openpyxl import load_workbook

    workbook = load_workbook(prediction, read_only=False, data_only=False)
    worksheet = workbook.active
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    positions = {
        column: header.index(column) + 1
        for column in QUEUE_ONLY_PREDICTION_COLUMNS
        if column in header
    }
    removed = [
        column for column in QUEUE_ONLY_PREDICTION_COLUMNS if column in positions
    ]
    rows = max(worksheet.max_row - 1, 0)
    temporary = prediction.with_suffix(f".tmp.{os.getpid()}.xlsx")
    try:
        for position in sorted(positions.values(), reverse=True):
            worksheet.delete_cols(position)
        if removed:
            workbook.save(temporary)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()
    if removed:
        temporary.replace(prediction)
    return {
        "contract": EVALUATOR_INPUT_FILTER_CONTRACT,
        "removed_columns": removed,
        "rows": rows,
        "evaluator_input_sha256": _sha256(prediction),
    }


_CHARTQAPRO_FINAL_MARKER = re.compile(
    r"\bthe[ \t]+answer[ \t]+is\b[ \t]*:?[ \t]*",
    re.IGNORECASE,
)
_CHARTQAPRO_ANSWER_TAG = re.compile(
    r"<answer>\s*(.*?)\s*</answer>", re.IGNORECASE | re.DOTALL
)
_BALANCED_MARKDOWN = ("**", "__", "~~", "`", "*", "_")


def _chartqapro_final_sentence(value: Any) -> str | None:
    text = str(value or "")
    matches = list(_CHARTQAPRO_FINAL_MARKER.finditer(text))
    if not matches:
        return None
    answer = text[matches[-1].end() :].splitlines()[0].strip()
    if not answer:
        return None
    previous = None
    while answer and answer != previous:
        previous = answer
        answer = answer.strip().rstrip(".?!。！？").strip()
        for marker in _BALANCED_MARKDOWN:
            if (
                answer.startswith(marker)
                and answer.endswith(marker)
                and len(answer) > 2 * len(marker)
            ):
                answer = answer[len(marker) : -len(marker)].strip()
                break
    return answer or None


def _chartqapro_last_boxed(text: str) -> str | None:
    last: str | None = None
    start = 0
    while True:
        marker = text.find("\\boxed{", start)
        if marker < 0:
            return last
        content_start = marker + len("\\boxed{")
        depth = 1
        pos = content_start
        while pos < len(text) and depth:
            if text[pos] == "{":
                depth += 1
            elif text[pos] == "}":
                depth -= 1
            pos += 1
        if depth:
            return last
        last = text[content_start : pos - 1].strip()
        start = pos


def _chartqapro_final_answer(value: Any) -> str | None:
    text = str(value or "")
    answer_blocks = list(_CHARTQAPRO_ANSWER_TAG.finditer(text))
    if answer_blocks:
        if len(answer_blocks) != 1:
            return None
        answer = answer_blocks[0].group(1).strip()
        if not answer:
            return None
        boxed = _chartqapro_last_boxed(answer)
        if boxed is not None:
            return boxed or None
        marked = _chartqapro_final_sentence(answer)
        return marked if marked is not None else answer
    return _chartqapro_final_sentence(text)


def _adapt_chartqapro_prediction(
    prediction: Path,
    table_loader: Callable[[str], Any],
) -> dict[str, Any]:
    data = table_loader(str(prediction)).copy()
    if "prediction" not in data:
        raise KeyError(f"ChartQAPro workbook lacks prediction column: {prediction}")
    original = data["prediction"].copy()
    data["raw_prediction"] = original
    changed = 0
    unresolved = 0
    adapted: list[Any] = []
    for raw in original:
        extracted = _chartqapro_final_answer(raw)
        if extracted is None:
            unresolved += 1
            adapted.append(raw)
        else:
            changed += int(str(extracted) != str(raw))
            adapted.append(extracted)
    data["prediction"] = adapted
    data.to_excel(prediction, index=False)
    return {
        "contract": CHARTQAPRO_ADAPTER_CONTRACT,
        "changed_rows": changed,
        "unresolved_rows": unresolved,
        "rows": int(len(data)),
        "adapted_prediction_sha256": _sha256(prediction),
    }


def _adapt_phyx_option_prediction(
    prediction: Path,
    table_loader: Callable[[str], Any],
) -> dict[str, Any]:
    sys.path.insert(0, str(SCRIPTS_ROOT))
    from run_external_benchmark_score_queue import _extract_option_letter

    data = table_loader(str(prediction)).copy()
    if "prediction" not in data:
        raise KeyError(f"PhyX workbook lacks prediction column: {prediction}")
    original = data["prediction"].copy()
    data["raw_prediction"] = original
    extracted = [_extract_option_letter(value, choices="ABCD") for value in original]
    data["prediction"] = [
        f"Answer: {value}" if value else "UNRESOLVED" for value in extracted
    ]
    data.to_excel(prediction, index=False)
    resolved = sum(bool(value) for value in extracted)
    return {
        "contract": PHYX_OPTION_ADAPTER_CONTRACT,
        "extractor": "run_external_benchmark_score_queue._extract_option_letter",
        "resolved_rows": resolved,
        "unresolved_rows": len(extracted) - resolved,
        "rows": int(len(data)),
        "adapted_prediction_sha256": _sha256(prediction),
    }


def _treebench_final_boxed_option(value: Any) -> str | None:
    text = str(value or "")
    answer_blocks = list(_CHARTQAPRO_ANSWER_TAG.finditer(text))
    if len(answer_blocks) != 1:
        return None
    answer = answer_blocks[0].group(1).strip()
    if not answer:
        return None
    match = re.search(
        r"\\boxed\s*\{\s*([A-E])\s*\}\s*[.!?。！？]*\s*$",
        answer,
        re.IGNORECASE,
    )
    return match.group(1).upper() if match else None


def _adapt_treebench_prediction(
    prediction: Path,
    table_loader: Callable[[str], Any],
) -> dict[str, Any]:
    data = table_loader(str(prediction)).copy()
    if "prediction" not in data:
        raise KeyError(f"TreeBench workbook lacks prediction column: {prediction}")
    original = (
        data["raw_prediction"].copy()
        if "raw_prediction" in data
        else data["prediction"].copy()
    )
    data["raw_prediction"] = original
    changed = 0
    adapted: list[Any] = []
    for raw in original:
        extracted = _treebench_final_boxed_option(raw)
        if extracted is None:
            adapted.append(raw)
        else:
            changed += int(str(extracted) != str(raw))
            adapted.append(extracted)
    data["prediction"] = adapted
    data.to_excel(prediction, index=False)
    return {
        "contract": TREEBENCH_OPTION_ADAPTER_CONTRACT,
        "changed_rows": changed,
        "rows": int(len(data)),
        "adapted_prediction_sha256": _sha256(prediction),
    }


def _adapt_explicit_mcq_prediction(
    prediction: Path,
    table_loader: Callable[[str], Any],
    *,
    dataset_name: str,
    contract: str,
) -> dict[str, Any]:
    """Normalize explicit final MCQ options before official exact matching.

    The benchmark is scored by the stock ``ImageMCQDataset.evaluate`` path.
    This adapter only removes model-specific answer wrappers (for example
    ``<answer>...\\boxed{B}</answer>``); rows without an explicit final option
    remain unchanged for VLMEvalKit's own matcher.
    """

    sys.path.insert(0, str(SCRIPTS_ROOT))
    sys.path.insert(0, str(EXTENSIONS_ROOT))
    from run_external_benchmark_score_queue import _extract_final_mcq_option
    from trace_eval_answer_parsing import extract_unambiguous_abcd
    import pandas as pd

    data = table_loader(str(prediction)).copy()
    if "prediction" not in data:
        raise KeyError(f"{dataset_name} workbook lacks prediction column: {prediction}")
    original = (
        data["raw_prediction"].copy()
        if "raw_prediction" in data
        else data["prediction"].copy()
    )
    data["raw_prediction"] = original
    adapted: list[Any] = []
    methods: dict[str, int] = {}
    resolved = 0
    for (_, row), raw in zip(data.iterrows(), original):
        choices = [
            label
            for label in "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
            if label in data.columns and not pd.isna(row[label])
        ]
        option = extract_unambiguous_abcd(raw)
        method = "explicit_wrapper"
        if option == "Z" or option not in choices:
            option, method = _extract_final_mcq_option(raw, choices)
        if option != "Z" and option in choices:
            adapted.append(f"Answer: {option}")
            resolved += 1
        else:
            adapted.append(raw)
            method = "official_matcher_fallback"
        methods[method] = methods.get(method, 0) + 1
    data["prediction"] = adapted
    data.to_excel(prediction, index=False)
    return {
        "contract": contract,
        "extractor": (
            "trace_eval_answer_parsing.extract_unambiguous_abcd + "
            "run_external_benchmark_score_queue._extract_final_mcq_option"
        ),
        "resolved_rows": resolved,
        "unresolved_rows": int(len(data)) - resolved,
        "methods": methods,
        "rows": int(len(data)),
        "adapted_prediction_sha256": _sha256(prediction),
    }


def _adapt_realworldqa_prediction(
    prediction: Path,
    table_loader: Callable[[str], Any],
) -> dict[str, Any]:
    return _adapt_explicit_mcq_prediction(
        prediction,
        table_loader,
        dataset_name="RealWorldQA",
        contract=REALWORLDQA_OPTION_ADAPTER_CONTRACT,
    )


def run_saved_score(
    *,
    benchmark_key: str | None,
    dataset_alias: str,
    prediction_xlsx: Path,
    output_dir: Path,
    model: str,
    model_slug: str,
    run_name: str,
    dataset_kwargs: dict[str, Any],
    judge_kwargs: dict[str, Any],
    primary_metric: str | None,
    primary_value_scale: str,
    vlmeval_root: Path,
    dataset_builder: Callable[..., Any] | None = None,
    table_loader: Callable[[str], Any] | None = None,
    flatten_metrics: Callable[[Any], dict[str, Any]] | None = None,
) -> dict[str, Any]:
    benchmark_key, dataset_alias, _ = _resolve_benchmark(benchmark_key, dataset_alias)
    prediction_xlsx = prediction_xlsx.expanduser().resolve()
    output_dir = output_dir.expanduser().resolve()
    vlmeval_root = vlmeval_root.expanduser().resolve()
    if prediction_xlsx.suffix.lower() != ".xlsx" or not prediction_xlsx.is_file():
        raise FileNotFoundError(f"Prediction XLSX does not exist: {prediction_xlsx}")
    if not vlmeval_root.is_dir():
        raise FileNotFoundError(f"VLMEvalKit checkout does not exist: {vlmeval_root}")
    commit = _git(vlmeval_root, "rev-parse", "HEAD")
    if commit != PINNED_VLMEVALKIT_COMMIT:
        raise RuntimeError(
            f"VLMEvalKit must be pinned to {PINNED_VLMEVALKIT_COMMIT}, found {commit}"
        )

    source_prediction_sha256 = _sha256(prediction_xlsx)
    sys.path.insert(0, str(vlmeval_root))
    using_default_builder = dataset_builder is None
    if dataset_builder is None or table_loader is None or flatten_metrics is None:
        from vlmeval.dataset import build_dataset
        from vlmeval.smp import load
        from vlmeval.smp.status_report import flatten_summary_metrics

        dataset_builder = dataset_builder or build_dataset
        table_loader = table_loader or load
        flatten_metrics = flatten_metrics or flatten_summary_metrics

    dataset_override: dict[str, Any] | None = None
    if benchmark_key == "erqa" and using_default_builder:
        from vlmeval.dataset.erqabench import ERQABench

        dataset = ERQABench(dataset=dataset_alias, **dict(dataset_kwargs))
        dataset_override = {
            "reason": "pinned registry selects ERQADataset whose evaluate references absent category metadata",
            "class": "vlmeval.dataset.erqabench.ERQABench",
        }
    else:
        dataset = dataset_builder(dataset_alias, **dict(dataset_kwargs))
    if dataset is None:
        raise RuntimeError(f"VLMEvalKit could not build dataset {dataset_alias!r}")
    official_name = str(getattr(dataset, "dataset_name", dataset_alias))
    if official_name != dataset_alias:
        raise ValueError(
            f"VLMEvalKit built {official_name!r} for alias {dataset_alias!r}"
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    source_columns = _xlsx_header(prediction_xlsx)
    has_queue_only_columns = any(
        column in source_columns for column in QUEUE_ONLY_PREDICTION_COLUMNS
    )
    prediction = _stage_prediction(
        prediction_xlsx,
        output_dir,
        official_name,
        replace=(
            benchmark_key
            in {
                "chartqapro",
                "phyx_mini_mc",
                "realworldqa",
                "treebench",
            }
            or has_queue_only_columns
        ),
    )
    evaluator_input_filter = _filter_evaluator_input(prediction)
    prediction_adapter: dict[str, Any] | None = None
    if benchmark_key == "chartqapro":
        prediction_adapter = _adapt_chartqapro_prediction(prediction, table_loader)
    elif benchmark_key == "phyx_mini_mc":
        prediction_adapter = _adapt_phyx_option_prediction(prediction, table_loader)
    elif benchmark_key == "treebench":
        prediction_adapter = _adapt_treebench_prediction(prediction, table_loader)
    elif benchmark_key == "realworldqa":
        prediction_adapter = _adapt_realworldqa_prediction(prediction, table_loader)

    # The only scoring call in this wrapper.
    result = dataset.evaluate(str(prediction), **dict(judge_kwargs))
    score, primary = _primary_score(
        dataset,
        result,
        flatten_metrics,
        primary_metric,
        primary_value_scale,
    )
    evaluate_source_raw = inspect.getsourcefile(dataset.evaluate)
    evaluate_source = (
        Path(evaluate_source_raw).resolve() if evaluate_source_raw else None
    )
    scores_path = output_dir / "scores.json"
    official_outputs = sorted(
        str(path.resolve())
        for path in output_dir.rglob("*")
        if path.is_file()
        and path.resolve() not in {prediction.resolve(), scores_path.resolve()}
    )
    status = _git(vlmeval_root, "status", "--short", "--untracked-files=normal")
    summary = {
        "dataset": official_name,
        "benchmark_key": benchmark_key,
        "model": model,
        "model_slug": model_slug,
        "run_name": run_name,
        "rows": int(len(table_loader(str(prediction)))),
        "score": score,
        "scores": _normalize_result(result),
        "primary_metric": primary,
        "artifacts": {
            "source_prediction_table": str(prediction_xlsx),
            "prediction_table": str(prediction),
            "official_outputs": official_outputs,
        },
        "provenance": {
            "contract": CONTRACT,
            "source_prediction_sha256": source_prediction_sha256,
            "prediction_sha256": _sha256(prediction),
            "evaluator_input_filter": evaluator_input_filter,
            "prediction_adapter": prediction_adapter,
            "dataset_override": dataset_override,
            "dataset_kwargs": _redact(dataset_kwargs),
            "judge_kwargs": _redact(judge_kwargs),
            "vlmevalkit_commit": commit,
            "vlmevalkit_worktree_dirty": bool(status),
            "dataset_class": f"{dataset.__class__.__module__}.{dataset.__class__.__qualname__}",
            "evaluate_source": str(evaluate_source) if evaluate_source else None,
            "evaluate_source_sha256": (
                _sha256(evaluate_source)
                if evaluate_source and evaluate_source.is_file()
                else None
            ),
            "evaluated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        },
    }
    _write_json(scores_path, summary)
    return summary


def _resolve_benchmark(key: str | None, alias: str | None) -> tuple[str, str, str]:
    if not key:
        raise ValueError("--benchmark-key is required")

    sys.path.insert(0, str(SCRIPTS_ROOT))
    from benchmark_queue_lib import spec_by_key
    from trace_eval_suite import load_trace_eval_suite

    spec = spec_by_key(key)
    suite = load_trace_eval_suite()
    if key not in suite.routes["official_vlmevalkit"]:
        route = next(item.route for item in suite.benchmarks if item.key == key)
        raise ValueError(
            f"Benchmark {key!r} uses the {route!r} route, " "not 'official_vlmevalkit'"
        )
    if alias is not None and alias != spec.alias:
        raise ValueError(f"Benchmark {key!r} uses alias {spec.alias!r}, not {alias!r}")
    return key, spec.alias, spec.run_name


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Score an existing prediction XLSX with pinned VLMEvalKit dataset.evaluate."
    )
    parser.add_argument("--benchmark-key", required=True)
    parser.add_argument("--dataset-alias")
    parser.add_argument("--prediction-xlsx", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--model", required=True)
    parser.add_argument("--model-slug", required=True)
    parser.add_argument("--run-name")
    parser.add_argument("--dataset-kwargs-json", default="{}")
    parser.add_argument("--judge-kwargs-json", default="{}")
    parser.add_argument("--primary-metric")
    parser.add_argument(
        "--primary-value-scale",
        choices=("auto", "percent", "fraction"),
        default="auto",
    )
    parser.add_argument(
        "--vlmeval-root",
        type=Path,
        default=Path(os.environ.get("VLMEVAL_ROOT", str(DEFAULT_VLMEVAL_ROOT))),
    )
    args = parser.parse_args()

    try:
        key, alias, default_run_name = _resolve_benchmark(
            args.benchmark_key, args.dataset_alias
        )
        summary = run_saved_score(
            benchmark_key=key,
            dataset_alias=alias,
            prediction_xlsx=args.prediction_xlsx,
            output_dir=args.output_dir,
            model=args.model,
            model_slug=args.model_slug,
            run_name=args.run_name or default_run_name,
            dataset_kwargs=_parse_json_object(
                args.dataset_kwargs_json, "--dataset-kwargs-json"
            ),
            judge_kwargs=_parse_json_object(
                args.judge_kwargs_json, "--judge-kwargs-json"
            ),
            primary_metric=args.primary_metric,
            primary_value_scale=args.primary_value_scale,
            vlmeval_root=args.vlmeval_root,
        )
    except Exception as error:
        parser.exit(1, f"error: {error}\n")
    print(json.dumps(summary, indent=2, ensure_ascii=False, default=_json_default))


if __name__ == "__main__":
    main()
