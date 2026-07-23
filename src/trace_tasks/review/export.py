"""Portable exports for local Trace task-review state."""

from __future__ import annotations

from datetime import datetime, timezone
import json
from pathlib import Path
import re
import tempfile
from typing import Any, Mapping

from .calibration import discover_calibration_summaries
from .app.index import build_review_index
from .app.store import ReviewStore


def export_review_report(
    *,
    review_root: Path | str,
    database_path: Path | str,
    output_path: Path | str,
    output_format: str = "json",
) -> Path:
    """Export materialization status and local review decisions."""

    root = Path(review_root).expanduser().resolve()
    destination = Path(output_path).expanduser().resolve()
    normalized_format = _resolve_format(output_format, destination)
    index = build_review_index(root)
    review_state = ReviewStore(database_path).export_snapshot()
    calibration_summaries, calibration_errors = discover_calibration_summaries(root)
    index_errors = [_portable_error(value) for value in index.errors]
    calibration_errors = [_portable_error(value) for value in calibration_errors]
    report = {
        "schema": "trace-review-report-v1",
        "created_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "summary": {
            "domain_count": len(index.domains),
            "scene_count": len(index.scenes),
            "task_count": len(index.tasks),
            "materialized_task_count": index.materialized_task_count,
            "sample_count": index.sample_count,
            "issue_count": len(review_state["issues"]),
            "audit_count": len(review_state["task_audits"]),
            "asset_review_count": len(review_state["asset_reviews"]),
            "calibration_report_count": len(calibration_summaries),
        },
        "tasks": [
            {
                "domain": task.domain,
                "scene_id": task.scene_id,
                "task_id": task.task_id,
                "materialized": task.materialized,
                "sample_count": task.sample_count,
                "query_counts": dict(sorted(task.query_counts.items())),
                "manifest": task.manifest_rel_path,
                "recipe_id": task.recipe_id,
                "recipe_digest": task.recipe_digest,
                "source_provenance": dict(task.source_provenance),
                "recipe_provenance": dict(task.recipe_provenance),
            }
            for task in sorted(
                index.tasks.values(),
                key=lambda task: (task.domain, task.scene_id, task.task_id),
            )
        ],
        **review_state,
        "calibration_summaries": calibration_summaries,
        "calibration_errors": calibration_errors,
        "index_errors": index_errors,
    }
    if normalized_format == "json":
        _atomic_text_write(
            destination, json.dumps(report, indent=2, sort_keys=True) + "\n"
        )
    elif normalized_format == "jsonl":
        records: list[dict[str, Any]] = [
            {"record_type": "summary", **report["summary"]}
        ]
        for key in (
            "tasks",
            "issues",
            "task_audits",
            "asset_reviews",
            "calibration_summaries",
        ):
            record_type = {
                "tasks": "task",
                "issues": "issue",
                "task_audits": "task_audit",
                "asset_reviews": "asset_review",
                "calibration_summaries": "calibration_summary",
            }[key]
            records.extend(
                {"record_type": record_type, **record} for record in report[key]
            )
        records.extend(
            {"record_type": "index_error", "error": value}
            for value in report["index_errors"]
        )
        records.extend(
            {"record_type": "calibration_error", "error": value}
            for value in report["calibration_errors"]
        )
        body = "".join(
            json.dumps(record, sort_keys=True, ensure_ascii=False) + "\n"
            for record in records
        )
        _atomic_text_write(destination, body)
    else:
        _write_workbook(destination, report)
    return destination


def _resolve_format(value: str, path: Path) -> str:
    normalized = str(value).strip().lower()
    if normalized == "auto":
        normalized = {
            ".json": "json",
            ".jsonl": "jsonl",
            ".xlsx": "xlsx",
        }.get(path.suffix.lower(), "")
    if normalized not in {"json", "jsonl", "xlsx"}:
        raise ValueError(
            "output_format must be one of {'json', 'jsonl', 'xlsx', 'auto'}"
        )
    return normalized


def _write_workbook(path: Path, report: Mapping[str, Any]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - review extra installs openpyxl.
        raise RuntimeError(
            "xlsx export requires the Trace review optional dependencies"
        ) from exc
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _append_xlsx_row(summary_sheet, ["Field", "Value"])
    for key, value in report["summary"].items():
        _append_xlsx_row(summary_sheet, [key, value])

    task_sheet = workbook.create_sheet("Tasks")
    task_headers = [
        "domain",
        "scene_id",
        "task_id",
        "materialized",
        "sample_count",
        "query_counts",
        "manifest",
        "recipe_id",
        "recipe_digest",
        "source_provenance",
        "recipe_provenance",
    ]
    _append_xlsx_row(task_sheet, task_headers)
    for task in report["tasks"]:
        _append_xlsx_row(
            task_sheet,
            [
                (
                    task.get(header)
                    if header
                    not in {
                        "query_counts",
                        "source_provenance",
                        "recipe_provenance",
                    }
                    else json.dumps(task.get(header, {}), sort_keys=True)
                )
                for header in task_headers
            ],
        )

    audit_sheet = workbook.create_sheet("Task audits")
    audits = list(report["task_audits"])
    audit_headers = list(audits[0].keys()) if audits else ["task_id"]
    _append_xlsx_row(audit_sheet, audit_headers)
    for audit in audits:
        _append_xlsx_row(audit_sheet, [audit.get(header) for header in audit_headers])

    issue_sheet = workbook.create_sheet("Issues")
    issues = list(report["issues"])
    issue_headers = [
        "id",
        "status",
        "severity",
        "category",
        "domain",
        "scene_id",
        "task_id",
        "sample_uid",
        "recipe_digest",
        "sample_semantic_hash",
        "title",
        "body",
        "author",
        "created_at",
        "updated_at",
        "comments",
    ]
    _append_xlsx_row(issue_sheet, issue_headers)
    for issue in issues:
        _append_xlsx_row(
            issue_sheet,
            [
                (
                    json.dumps(issue.get(header, []), sort_keys=True)
                    if header == "comments"
                    else issue.get(header)
                )
                for header in issue_headers
            ],
        )

    asset_sheet = workbook.create_sheet("Asset reviews")
    asset_reviews = list(report["asset_reviews"])
    asset_headers = (
        list(asset_reviews[0].keys())
        if asset_reviews
        else ["asset_id", "kind", "decision"]
    )
    _append_xlsx_row(asset_sheet, asset_headers)
    for review in asset_reviews:
        _append_xlsx_row(asset_sheet, [review.get(header) for header in asset_headers])

    calibration_sheet = workbook.create_sheet("Calibration")
    calibration_headers = [
        "source",
        "created_at",
        "model",
        "task_ids",
        "sample_count",
        "summary",
        "settings",
        "provenance",
    ]
    _append_xlsx_row(calibration_sheet, calibration_headers)
    for calibration in report["calibration_summaries"]:
        _append_xlsx_row(
            calibration_sheet,
            [
                (
                    json.dumps(calibration.get(header), sort_keys=True)
                    if header in {"task_ids", "summary", "settings", "provenance"}
                    else calibration.get(header)
                )
                for header in calibration_headers
            ],
        )

    error_sheet = workbook.create_sheet("Errors")
    _append_xlsx_row(error_sheet, ["kind", "error"])
    for value in report["index_errors"]:
        _append_xlsx_row(error_sheet, ["index", value])
    for value in report["calibration_errors"]:
        _append_xlsx_row(error_sheet, ["calibration", value])

    with tempfile.NamedTemporaryFile(
        suffix=".xlsx", prefix=f".{path.name}.", dir=path.parent, delete=False
    ) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        temporary.replace(path)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise


def _atomic_text_write(path: Path, body: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w", encoding="utf-8", dir=path.parent, prefix=f".{path.name}.", delete=False
    ) as handle:
        temporary = Path(handle.name)
        handle.write(body)
    try:
        temporary.replace(path)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise


def _append_xlsx_row(sheet: Any, values: list[Any]) -> None:
    sheet.append([_xlsx_cell(value) for value in values])


def _xlsx_cell(value: Any) -> Any:
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


_ABSOLUTE_PATH_RE = re.compile(
    r"(?<![A-Za-z0-9_.-])(?:/[A-Za-z0-9_.~:@%+\-]+)+|"
    r"(?<![A-Za-z0-9_.-])[A-Za-z]:\\[^\s]+"
)


def _portable_error(value: Any) -> str:
    text = str(value).strip().replace("\r", " ").replace("\n", " ")
    return _ABSOLUTE_PATH_RE.sub("<path>", text)[:500] or "operation failed"


__all__ = ["export_review_report"]
